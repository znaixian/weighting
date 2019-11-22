# -*- coding: utf-8 -*-
"""
Created on Fri Nov 22 14:03:12 2019

@author: ncarucci
"""

from django.shortcuts import render
import openpyxl
import pandas as pd
import os
import requests

def Index(request):
    return render(request, 'weighting/index.html', {"data_list": []})

#def index(request):
#    r = requests.get('http://httpbin.org/status/418')
#    print(r.text)
#    return HttpResponse('<pre>' + r.text + '</pre>')

def Capping(df, threshold):
    while (df.weight > threshold).any():
        largest = float(df.weight.nlargest(1)) 
        df['weight_1'] = 0
        df.loc[df.weight == threshold, 'weight_1'] = threshold
        # df['weight_1'][df.weight == threshold] = threshold
        num = len(df[df.weight == largest]) 
        df.loc[df.weight == largest, 'weight_1'] = threshold
        # df['weight_1'][df.weight == largest] = threshold
        dist = (largest - threshold)*num
        total = df.weight[(df.weight_1 == 0)].sum()
        df.loc[df.weight_1 == 0, 'weight_1'] = df.weight + dist*(df.weight/total)
        # df['weight_1'][df.weight_1 == 0] = df.weight + dist*(df.weight/total)
        del df['weight']
        df.rename(columns={'weight_1': 'weight'}, inplace=True)
    return df

def Upload(request):
    excel_data_list = []
    error_dict = {}
    file_path = ""
    if request.method == 'POST' and request.FILES:
        f = request.FILES['input_data']
        t = request.POST['threshold']
        if t:
            pd.set_option("display.precision", 16)
            wb = openpyxl.load_workbook(f)
            worksheet = wb["Sheet1"]

            excel_data = list()
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)
                
            excel_data_list = []
            for data in excel_data[1:]:
                print data
                
                excel_data_dict = {}
                excel_data_dict['Date'] = data[0]
                excel_data_dict["sedol"] = data[1]
                excel_data_dict["float_market_cap"] = data[2]
                excel_data_dict["weight"] = float(data[3])*100 if data[3] != 'None' else 0
                excel_data_list.append(excel_data_dict)
            
            df = pd.DataFrame(excel_data_list)
            udates = df.Date.unique()
            
            tb = pd.DataFrame()
            for day in udates:
              print day
              dfile = df[df.Date == day]
              Capping(dfile, float(t))
              tb = tb.append(dfile)
    
            # Create Excel File
#            wr = openpyxl.Workbook() 
#            tb = wr.active

            tb.to_excel(os.path.join("static/files", request.FILES['input_data'].name))
            file_path = os.path.join("static/files/", request.FILES['input_data'].name)
            print(file_path)
        else:
            excel_data_list = []
            error_dict['file_error']="Please Upload File!"
            error_dict['threshold_error'] = "Please Add Threshold!"
    elif request.method == 'POST':
        error_dict = {}
        error_dict['threshold_error'] = ""
        t = request.POST['threshold']
        if not t:
            error_dict['threshold_error'] = "Please Add Threshold!"
        error_dict['file_error']="Please Upload File!"

    return render(request, 'weighting/index.html', {"data_list":excel_data_list, "error_dict": error_dict, "file_path":file_path})
